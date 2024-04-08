import time, serial, sys, os, cv2
from tkinter import Tk, Frame, Label, font, PhotoImage
from scipy import *
from numpy import array
from PIL import Image, ImageTk
from pyzbar.pyzbar import decode
import json
import locale
from datetime import datetime
import pandas as pd 
from openpyxl import load_workbook



alunos_entrada = []

locale.setlocale(locale.LC_ALL, "pt_BR")
diaPT = datetime.now().strftime('%d/%m/%Y')
today = datetime.now().strftime('%a')

camera_id = 0
delay = 1
window_name = 'OpenCV pyzbar'

alunos_dict = {}
alunos_img = {}
alunos_data = pd.read_csv('./Lista.csv').to_dict()

for index in range(len(alunos_data.get('nome'))):
    alunos_dict[alunos_data['nome'][index]] = alunos_data['dias'][index].split(',') 
    alunos_img[alunos_data['nome'][index]] = alunos_data['imagem'][index]

print(alunos_dict.values())
print(alunos_img.values())

mGui = Tk()
mGui.geometry('600x600+0+0')
mGui.configure(background="#3a4447")
mGui.state('zoomed')
camFrame = Frame(mGui, width=435, height=475)
camFrame.place(x=100, y=70)
infoFrame = Frame(mGui, width=675, height=475, bg='#667174')
infoFrame.place(x=500, y=70)


cap = cv2.VideoCapture(camera_id)
ret, frame = cap.read()

v1 = Label(camFrame, text="QrCode Video")
v1.place(x=-80, y=-2)

almoco = 'Aponte o QR CODE a câmera para leitura:'
v2 = Label(infoFrame, text=almoco, background='#da251c', foreground='white', font=('Arial', 20, 'bold'))
v2.place(x=59, y=20)
v2.config(borderwidth=2, relief="solid")

def dddd():
    ret, frame = cap.read()
    global almoco
    if ret:
        for d in decode(frame):
            try:
                s = d.data.decode()
                json_data = json.loads(s)
                nome_aluno = json_data['nome']
                dias_almoco = json_data['comida']
                almoco = f'{nome_aluno} NAO LIBERADO(A)'

                #PROCURAR SE TEM O NOME NO DIC, SE O TODAY TEM NA LISTA DO QRCODE E SE O A LISTA DO QRCODE É IGUAL A DA PLANILHA
                if nome_aluno in alunos_dict.keys() and today in dias_almoco and dias_almoco == alunos_dict[nome_aluno]:
                    
                    
                    #PROCURAR SE O ALUNO ESTÁ NA PLANILHA
                    wb = load_workbook("relacao.xlsx")
                    ws = wb.active
                    aluno_presente = False

                    for row in ws.iter_rows(values_only=True):
                        if nome_aluno in row and diaPT in row:
                            aluno_presente = True
                            break

                    #VERIFICAR SE O ALUNO JA PASSOU OU NAO
                    if not aluno_presente:
                        #ADD ALUNO NO EXCEL
                        # Carregar o arquivo do Excel
                        wb = load_workbook("relacao.xlsx")

                        # Obter a primeira planilha (índice 0)
                        ws = wb.worksheets[0]

                        # Adicionar novo aluno ao Excel
                        new_row_data = [[nome_aluno, today, diaPT]]
                        for row_data in new_row_data:
                            ws.append(row_data)

                        # Salvar o arquivo Excel
                        wb.save("relacao.xlsx")


                        #ALUNO LIBERADO E IMAGEM
                        almoco = f'{nome_aluno} LIBERADO(A)'
                        v2.config(foreground='white') 
                        imagem_aluno =  alunos_img[nome_aluno]
                        imagem = Image.open('fotos/' + imagem_aluno)
                        imagem.thumbnail((500, 500)) 
                        foto_aluno = ImageTk.PhotoImage(imagem)

                        label_foto_aluno = Label(mGui, image=foto_aluno)
                        label_foto_aluno.image = foto_aluno 
                        image_height = imagem.height 
                        label_foto_aluno.place(x=700, y=200)  

                        #CARREGAR QRCODE
                        time.sleep(2)
                    
                    else:
                        almoco = f'{nome_aluno} JA PASSOU'
                        imagem_aluno =  alunos_img[nome_aluno]
                        imagem = Image.open('fotos/' + imagem_aluno)
                        imagem.thumbnail((500, 500)) 
                        foto_aluno = ImageTk.PhotoImage(imagem)

                        label_foto_aluno = Label(mGui, image=foto_aluno)
                        label_foto_aluno.image = foto_aluno 
                        image_height = imagem.height 
                        label_foto_aluno.place(x=700, y=200) 
                              
                else:
                    almoco = f'{nome_aluno} NAO LIBERADO(A)!'

                    imagem = Image.open('fotos/error.jpg')
                    imagem.thumbnail((200, 200)) 
                    foto_aluno = ImageTk.PhotoImage(imagem)
                    label_foto_aluno = Label(mGui, image=foto_aluno)
                    label_foto_aluno.image = foto_aluno
                    image_height = imagem.height 
                    label_foto_aluno.place(x=700, y=200)  

                    v2.config(foreground='black')
                    
            except Exception as e:
                print(e)
                pass
    img = Image.fromarray(cv2.cvtColor(frame, cv2.COLOR_RGB2BGR))
    nimg = ImageTk.PhotoImage(image=img)
    
   

    image = Image.open("sesi.png")
    image.thumbnail((150, 150))
    photo = ImageTk.PhotoImage(image)


    image_label = Label(mGui, image=photo)
    image_label.image = photo 


    image_height = image.height 

    image_label.place(x=10, y=645 - image_height)  


    
    v1.n_img = nimg
    v1.configure(image=nimg)
    v2.config(text=almoco)
    
    mGui.after(10, dddd)
dddd()
mGui.mainloop()